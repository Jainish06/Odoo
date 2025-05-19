import binascii
import tempfile
import openpyxl

from odoo import api, fields, models
from odoo.exceptions import UserError


class MrpReplaceWizard(models.TransientModel):
    _name = 'mrp.replace.wizard'
    _description = 'Replace BOM'
    _rec_name = 'operation_type'

    operation_type = fields.Selection([('update_product_code', 'Update Product Code'), ('update_serial', 'Update Serial')], string='Operation Type', default='update_product_code')
    excel_file = fields.Binary(string='Upload Excel File')
    products_to_update_code_ids = fields.One2many('mrp.replace.products', 'mrp_replace_wizard_id', string='Products to Update Code')
    products_to_update_serial_ids = fields.One2many('mrp.update.serial', 'mrp_replace_wizard_id', string='Products to Update Code')

    def generate_replace_product_lines(self):
        try:
            fp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            fp.write(binascii.a2b_base64(self.excel_file))
            fp.seek(0)
        except:
            raise UserError("Invalid File!")

        try:
            wb = openpyxl.load_workbook(fp.name)
            ws = wb.active
            print(ws)
            if self.operation_type == 'update_product_code':
                update_code_list = []
                self.products_to_update_code_ids = [(5, 0, 0)]
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=True):
                    update_code_list.append((0, 0, {
                        'mo_num' : record[0],
                        'current_product_code' : record[1],
                        'new_product_code' : record[2]
                    }))
                self.products_to_update_code_ids = update_code_list
            else:
                update_serial_list = []
                self.products_to_update_serial_ids = [(5, 0, 0)]
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=True):
                    update_serial_list.append((0, 0, {
                        'mo_num': record[0],
                        'current_product_code': record[1],
                        'old_product_serial': record[2],
                        'new_product_serial' : record[3],
                    }))
                self.products_to_update_serial_ids = update_serial_list

        except Exception as e:
            print(e)
        return {
            'name': 'Replace Product',
            'view_mode': 'form',
            'res_model': 'mrp.replace.wizard',
            'res_id': self.id,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }

    def action_replace(self):
        pass

